"""Generate XLS report from Gemini VideoAnalysis.

Summary sheet   : metadata, character screen time table, punchy dialogues.
Character sheets: one per character — appearances, top clips.
"""
from __future__ import annotations

import os
import re
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from models.gemini_responses import VideoAnalysis

_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL    = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")


def _header_row(ws, row: int, cols: list[str]) -> None:
    for ci, name in enumerate(cols, 1):
        cell = ws.cell(row=row, column=ci, value=name)
        cell.fill, cell.font = _HEADER_FILL, _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 25


def _auto_width(ws) -> None:
    for col in ws.columns:
        w = max((len(str(c.value or "")) for c in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, 60)


def _summary_sheet(wb, analysis: VideoAnalysis,
                   url: str, title: str, duration: float) -> None:
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = title or "Video Analysis Report"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A2"] = f"Duration: {duration:.1f}s  |  Genre: {analysis.genre}"
    ws["A3"] = f"URL: {url}"
    ws["A4"] = analysis.visual_summary
    ws["A4"].alignment = Alignment(wrap_text=True)
    ws.merge_cells("A4:F4")
    ws.row_dimensions[4].height = 60

    _header_row(ws, 6, ["Rank", "Character", "Screen Time (s)",
                         "Screen Time (min)", "Scene Count", "% of Video"])
    for rank, char in enumerate(analysis.characters, 1):
        pct = char.total_screen_time_seconds / duration * 100 if duration else 0
        scene_count = sum(1 for s in analysis.scene_summaries
                          if char.character_id in s.characters_present)
        for ci, v in enumerate([
            rank, char.name,
            f"{char.total_screen_time_seconds:.2f}",
            f"{char.total_screen_time_seconds/60:.2f}",
            scene_count, f"{pct:.1f}%",
        ], 1):
            cell = ws.cell(row=6 + rank, column=ci, value=v)
            if rank % 2 == 0:
                cell.fill = _ALT_FILL
            cell.alignment = Alignment(horizontal="center")

    char_id_to_name = {c.character_id: c.name for c in analysis.characters}
    dlg_row = 6 + len(analysis.characters) + 3
    ws.cell(row=dlg_row, column=1, value="Top Punchy Dialogues").font = Font(bold=True, size=12)
    _header_row(ws, dlg_row + 1, ["Rank", "Speaker", "Dialogue", "Start (s)", "Reasoning"])
    for dlg in analysis.punchy_dialogues:
        speaker = char_id_to_name.get(dlg.speaker, dlg.speaker)
        for ci, v in enumerate(
            [dlg.rank, speaker, dlg.text, f"{dlg.start_sec:.1f}", dlg.reasoning], 1
        ):
            cell = ws.cell(row=dlg_row + 1 + dlg.rank, column=ci, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[dlg_row + 1 + dlg.rank].height = 60
    _auto_width(ws)


def _derive_appearances(char_id: str, scenes: list) -> list[tuple[float, float]]:
    """Group consecutive scenes where character appears into appearance blocks."""
    char_scenes = sorted(
        [(s.start_sec, s.end_sec) for s in scenes if char_id in s.characters_present]
    )
    if not char_scenes:
        return []
    blocks, cur_start, cur_end = [], char_scenes[0][0], char_scenes[0][1]
    for start, end in char_scenes[1:]:
        if start <= cur_end + 35:   # merge scenes within ~35s gap
            cur_end = max(cur_end, end)
        else:
            blocks.append((cur_start, cur_end))
            cur_start, cur_end = start, end
    blocks.append((cur_start, cur_end))
    return blocks


def _character_sheet(wb, char, rank: int, scenes: list) -> None:
    appearances = _derive_appearances(char.character_id, scenes)
    ws = wb.create_sheet(title=f"#{rank} {char.name}"[:31])
    ws["A1"] = f"Character: {char.name}"
    ws["A1"].font = Font(bold=True, size=13)
    ws["A2"] = (f"Screen Time: {char.total_screen_time_seconds:.2f}s  "
                f"({char.total_screen_time_seconds/60:.2f} min)")
    ws["A3"] = f"Appearances: {len(appearances)}"
    ws["A4"] = f"Description: {char.description}"
    for ref in ("A1:F1", "A2:F2", "A3:F3", "A4:F4"):
        ws.merge_cells(ref)

    _header_row(ws, 6, ["#", "Start (s)", "End (s)", "Duration (s)"])
    for i, (start, end) in enumerate(appearances, 1):
        dur = end - start
        for ci, v in enumerate([i, f"{start:.1f}", f"{end:.1f}", f"{dur:.1f}"], 1):
            cell = ws.cell(row=6 + i, column=ci, value=v)
            if i % 2 == 0:
                cell.fill = _ALT_FILL
            cell.alignment = Alignment(horizontal="center")

    clip_row = 6 + len(appearances) + 3
    ws.cell(row=clip_row, column=1, value="Top Recommended Clips").font = Font(bold=True, size=12)
    _header_row(ws, clip_row + 1,
                ["Clip #", "Title", "Start (s)", "End (s)", "Duration (s)", "Reasoning"])
    for j, clip in enumerate(char.top_clips, 1):
        dur = clip.end_sec - clip.start_sec
        for ci, v in enumerate(
            [j, clip.title, f"{clip.start_sec:.2f}", f"{clip.end_sec:.2f}",
             f"{dur:.2f}", clip.reasoning], 1
        ):
            cell = ws.cell(row=clip_row + 1 + j, column=ci, value=v)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.row_dimensions[clip_row + 1 + j].height = 50
    _auto_width(ws)


def generate_report(analysis: VideoAnalysis, output_dir: str,
                    url: str, video_id: str, title: str, duration: float) -> str:
    wb = openpyxl.Workbook()
    _summary_sheet(wb, analysis, url, title, duration)
    for rank, char in enumerate(analysis.characters, 1):
        _character_sheet(wb, char, rank, analysis.scene_summaries)
    safe_title = re.sub(r'[^\w\s-]', '', title)
    safe_title = re.sub(r'[\s-]+', '_', safe_title).strip('_')[:60] or video_id
    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename   = f"report_{safe_title}_{timestamp}.xlsx"

    Path(output_dir).mkdir(parents=True, exist_ok=True)
    path = str(Path(output_dir) / filename)
    wb.save(path)
    return path
